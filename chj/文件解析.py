import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import re
import os
import subprocess
import platform
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


'''
上传一个excel、一个txt文件，实现userId过滤
'''
def process_files():
    try:
        # 让用户选择Excel文件
        excel_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not excel_path:
            return

        # 让用户选择txt文件
        txt_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if not txt_path:
            return

        # 读取Excel文件
        df = pd.read_excel(excel_path, dtype={'UserId': str})

        # 读取txt文件
        with open(txt_path, 'r',encoding='utf-8') as file:
            insert_lines = file.readlines()

        # 提取txt文件中的UserId
        txt_user_ids = set()
        for line in insert_lines:
            match = re.search(r"INSERT INTO .* VALUES\s*\(NEWID\(\), '([^']+)'", line)
            if match:
                user_id = match.group(1)
                txt_user_ids.add(user_id)

        # Excel数据去重
        df_unique = df.drop_duplicates(subset='UserId')

        # 与txt文件中的UserId进行比较去重
        unique_user_ids = df_unique[~df_unique['UserId'].astype(str).isin(txt_user_ids)].copy()
        duplicate_user_ids = df_unique[df_unique['UserId'].astype(str).isin(txt_user_ids)].copy()

        # 为unique_user_ids添加insert语句列
        def create_insert_statement(row):
            return f"INSERT INTO UserAgentConfig (Id, UserId, AgentMainDataId, Description, CreateTime, CreateBy, UpdateTime, UpdateBy, IsDeleted, Version) VALUES (NEWID(), '{row['UserId']}', '{row['AgentMainDataId']}', '{row['Description']}', GETDATE(), 'Import', NULL, NULL, '0', '1');"

        unique_user_ids.loc[:, 'InsertStatement'] = unique_user_ids.apply(create_insert_statement, axis=1)

        # 将插入语句保存到txt文件
        output_txt_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
        if not output_txt_path:
            return

        with open(output_txt_path, 'w',encoding='utf-8') as file:
            for insert_statement in unique_user_ids['InsertStatement']:
                file.write(insert_statement + '\n')

        # 标记重复的UserId为红色
        wb = load_workbook(excel_path)
        ws = wb.active

        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        for idx, row in duplicate_user_ids.iterrows():
            for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                if cell[0].value == row['UserId']:
                    for c in cell:
                        c.fill = red_fill

        # 自动生成带有标记的Excel文件名
        output_excel_path = os.path.join(os.path.dirname(excel_path), "DuplicateDataIsMarked.xlsx")

        wb.save(output_excel_path)

        # 提示成功信息
        messagebox.showinfo("Success", "文件已成功处理并保存.")

        # 自动打开生成的Excel文件
        if platform.system() == 'Darwin':  # macOS
            subprocess.call(('open', output_txt_path))
        elif platform.system() == 'Linux':  # Linux
            subprocess.call(('xdg-open', output_txt_path))
        else:  # Windows
            os.startfile(output_txt_path)

        # 关闭 root
        root.destroy()

    except Exception as e:
        # 提示错误信息
        messagebox.showerror("Error", str(e))


# 创建主窗口
root = tk.Tk()
root.title("文件处理器")

# 设置窗口大小和位置
window_width = 400
window_height = 300

# 获取屏幕尺寸以计算居中的位置
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

position_top = int(screen_height / 2 - window_height / 2)
position_right = int(screen_width / 2 - window_width / 2)

root.geometry(f'{window_width}x{window_height}+{position_right}+{position_top}')

# 创建说明标签
instructions = (
    "此工具将处理您上传的Excel和txt文件。\n"
    "1. 请您上传一个一个excel、一个txt文件。\n"
    "2. Excel文件中的UserId将与txt文件中的UserId进行比较。\n"
    "3. 未在txt文件中找到的UserId将生成insert语句并保存到新的txt文件中。\n"
    "4. 重复的UserId将在Excel中标记为红色。\n"
    "5. 生成的Excel文件名为DuplicateDataIsMarked.xlsx。\n"
)

label = tk.Label(root, text=instructions, wraplength=350, justify="left")
label.pack(pady=10)

# 创建按钮并设置大小和位置
process_button = tk.Button(root, text="上传文件并处理", command=process_files, width=20, height=2)
process_button.pack(expand=True)

# 运行主循环
root.mainloop()
